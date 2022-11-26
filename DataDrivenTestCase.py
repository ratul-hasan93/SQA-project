import time
import XLUrh
import openpyxl
from selenium import webdriver
from openpyxl.workbook import workbook
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
#from selenium.webdriver.chrome.service import Service

# code start
driver = webdriver.Chrome(executable_path="C://Drivers/chromedriver_win32/chromedriver.exe/")
driver.get("https://www.google.com/")
driver.maximize_window()
time.sleep(2)

path = "C://Users/Ratul Hasan/Desktop/Selenium/keyword.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active
rows = sheet.max_row
clos = sheet.max_column
# print(rows,clos)
time.sleep(5)

rows = XLUrh.getRowCount(path, 'Saturday')
# cols = XLUrh.getColumnCount(path, 'Saturday')

for r in range(2, rows+1):
     LongestOption = XLUrh.readData(path, "Saturday", r, 1)
     driver.find_element(By.NAME,'q').send_keys(LongestOption)
     time.sleep(5)
     driver.find_element(By.NAME,"btnK").click()
     driver.send_keys(keys.ENTER)
     time.sleep(5)
 

#     if driver.title == "cricket - Google Search":
#         print("test is passed")
#         XLUrh.writeData(path, "Saturday", r, 3, "test passed")
#     else:
#         print("test is faild")
#         XLUrh.writeData(path,"Saturday", r, 3, "test faild")

#     driver.find_element("Search Google or type a URL").click()

